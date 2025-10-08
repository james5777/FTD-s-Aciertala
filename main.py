# ----------------------------------------------------------------------
# Autor: James Palacio
#        Jacobo Chica
#
# Fecha: 22/09/2025
#
# Descripci√≥n: El objetivo de esta automatizacion es definir desde los exportables del sportbook,
# cuales de los usuarios registrados son primeros depositos, teniendo en cuenta las variables
# ----------------------------------------------------------------------

import pandas as pd
import sqlite3
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import auto_descarga as ad


def reporte_diario_aciertala():
    ### ---------- Archivos a leer ---------- ###
    archivo_registros_excel = Path("Archivos/Diario-ACPE.xlsx")
    archivo_transacciones_csv = Path("Archivos/transactions.csv")

    ### ---------- Nombres de tablas en SQLite ---------- ###
    name_tabla_registros_excel = "tabla_registros"
    name_tabla_transacciones_csv = "tabla_transacciones"
    name_tabla_resultados = 'cruce_registros_transacciones'
    name_tabla_resumen = 'tabla_resumen_diario'

    ### ---------- Archivo base de datos ---------- ###
    archivo_base_datos = Path("Archivos/Archivos_base_datos/DataBase_aciertala.db")

    ### ---------- Nombres de columnas archivo excel (Registros) ---------- ###
    ID_usuario_excel = 'ID de usuario'
    Usuario_excel = 'Usuario'
    fecha_creacion_excel = 'Fecha de creaci√≥n'

    ### ---------- Nombres de columnas archivo CSV (transacciones) ---------- ###
    fecha_deposito_csv = 'Crear hora'
    ID_usuario_csv = 'ID de usuario'
    usuario_csv = 'Usuario'
    tipo_transaccion_csv = 'Tipo de transacci√≥n'
    ID_transaccion_csv = 'ID de transacci√≥n'



    ### ---------- Conexi√≥n a SQLite ---------- ###
    conn = sqlite3.connect(archivo_base_datos)


    #####################################################################################################################################
    ########################################---------------BLOQUE PARA COMENTAR---------------###########################################
    #####################################################################################################################################
    ####----- En caso de necesitar actualizar algo en las tablas que se crean a partir de una consulta SQL, se comenta este bloque para que no se carguen mas datos a la base de datos (LO CUAL LA CORROMPERIA) -----#####

    ## ---------- Esta funcion se encarga de insertar la nueva data tanto de registros como de transacciones ---------- ###
    ## ---------- Importante tener en cuenta que para el excel, no es necesario subir solamente los nuevos registros ya que cuando, subimos el archivo, la funcion detecta por medio del usuario cuales datos estan duplicados y los omite, para los csv de transacciones si es necesario, seleccionar las fechas nuevas. en el sportbook seleccionar siempre la opcion "Ayer" exportarlo y subirlo ---------- ###

    def actualizar_tabla(df, nombre_tabla, conn, unique_cols):
        """
        Inserta solo filas nuevas en una tabla SQLite desde un DataFrame.
        """

        df_copy = df.copy() 
        
        # Convertir todas las columnas datetime a texto YYYY-MM-DD
        for col in df_copy.columns:
            if pd.api.types.is_datetime64_any_dtype(df_copy[col]) or df_copy[col].apply(lambda x: isinstance(x, pd.Timestamp)).any():
                df_copy.loc[:, col] = pd.to_datetime(df_copy[col], errors="coerce").dt.strftime("%Y-%m-%d")

        df_copy.head(0).to_sql(nombre_tabla, conn, if_exists="append", index=False)
        
        try:
            existentes = pd.read_sql(f"SELECT {', '.join(unique_cols)} FROM {nombre_tabla}", conn)
        except Exception:
            existentes = pd.DataFrame(columns=unique_cols)
        
        # Resto del c√≥digo...
        df_copy["_hash"] = df_copy[unique_cols].astype(str).agg("-".join, axis=1).astype(str)
        if not existentes.empty:
            existentes["_hash"] = existentes[unique_cols].astype(str).agg("-".join, axis=1).astype(str)
        else:
            existentes["_hash"] = []

        nuevos = df_copy[~df_copy["_hash"].isin(existentes["_hash"])].drop(columns=["_hash"])

        if not nuevos.empty:
            nuevos.to_sql(nombre_tabla, conn, if_exists="append", index=False)
            print(f"‚úÖ {len(nuevos)} filas nuevas insertadas en {nombre_tabla}")
        else:
            print(f"‚ÑπÔ∏è No hay filas nuevas en {nombre_tabla}")
        
        print(unique_cols)

    ### ---------- Leer Excel y CSV ---------- ###
    df_registros_excel = pd.read_excel(archivo_registros_excel, sheet_name='Historico')
    df_transacciones_csv = pd.read_csv(
        archivo_transacciones_csv,
        sep=",",
        quotechar='"',
        encoding="utf-8"
    )

    ### ---------- Normalizacion de columnas por fecha y por separadores de columnas ---------- ###
    df_transacciones_csv.columns = df_transacciones_csv.columns.str.strip().str.replace('"', '', regex=False)
    df_transacciones_csv[fecha_deposito_csv] = df_transacciones_csv[fecha_deposito_csv].str.split(' ').str[0]
    df_transacciones_csv[fecha_deposito_csv] = pd.to_datetime(df_transacciones_csv[fecha_deposito_csv]).dt.strftime('%Y-%m-%d')
    df_registros_excel[fecha_creacion_excel] = pd.to_datetime(df_registros_excel[fecha_creacion_excel]).dt.strftime('%Y-%m-%d')

    ### ---------- Se crea un nuevo dataframe unicamente con las columnas necesarias para el csv de transacciones y se normaliza a fecha sin hora ---------- ###
    col_necesarias_csv = [fecha_deposito_csv, ID_usuario_csv, usuario_csv, tipo_transaccion_csv, ID_transaccion_csv]
    df_final_transacciones = df_transacciones_csv[col_necesarias_csv]
    df_final_transacciones.loc[:, fecha_deposito_csv] = pd.to_datetime(df_final_transacciones[fecha_deposito_csv], errors="coerce").dt.normalize()

    ### ---------- Se crea un nuevo dataframe unicamente con las columnas necesarias para el excel de registros y se normaliza a fecha sin hora ---------- ###
    col_necesarias_excel = [fecha_creacion_excel, ID_usuario_excel, Usuario_excel]
    df_final_registros = df_registros_excel[col_necesarias_excel]
    df_final_registros.loc[:, fecha_creacion_excel] = pd.to_datetime(df_final_registros[fecha_creacion_excel], errors="coerce").dt.normalize()

    ### ---------- Actualizar tablas de registros y transacciones ---------- ###
    actualizar_tabla(df_final_registros, name_tabla_registros_excel, conn, unique_cols=[Usuario_excel])
    actualizar_tabla(df_final_transacciones, name_tabla_transacciones_csv, conn, unique_cols=[usuario_csv, fecha_deposito_csv, ID_transaccion_csv])

    print("üìå Tablas intermedias actualizadas. Ahora se realizar√° el cruce final.")

    #####################################################################################################################################
    ####################################---------------FIN DE BLOQUE PARA COMENTAR---------------########################################
    #####################################################################################################################################

    # ### ---------- Cruce con datos de la BD ---------- ###

    # Leer los datos completos desde las tablas de la BD
    reg = pd.read_sql(f"SELECT * FROM {name_tabla_registros_excel}", conn)
    txn = pd.read_sql(f"SELECT * FROM {name_tabla_transacciones_csv}", conn)

    # Asegurar tipos de datos correctos para el cruce
    reg.loc[:, Usuario_excel] = reg[Usuario_excel].fillna("").astype(str).str.strip().str.lower()
    txn.loc[:, usuario_csv] = txn[usuario_csv].fillna("").astype(str).str.strip().str.lower()
    reg.loc[:, fecha_creacion_excel] = pd.to_datetime(reg[fecha_creacion_excel], errors="coerce").dt.normalize()
    txn.loc[:, fecha_deposito_csv] = pd.to_datetime(txn[fecha_deposito_csv], errors="coerce").dt.normalize()

    # Filtrar SOLO dep√≥sitos
    mask_deposit = txn[tipo_transaccion_csv].fillna("").astype(str).str.lower().str.contains("deposit")
    txn_deposits = txn.loc[mask_deposit, [usuario_csv, fecha_deposito_csv, ID_transaccion_csv]].copy()
    txn_deposits = txn_deposits[txn_deposits[usuario_csv] != ""]

    # Obtener primer dep√≥sito
    idx = txn_deposits.groupby(usuario_csv)[fecha_deposito_csv].idxmin().dropna()
    first_deposits = txn_deposits.loc[idx].reset_index(drop=True)

    # Merge LEFT registros vs primeros dep√≥sitos
    df_resultado = reg.merge(
        first_deposits[[usuario_csv, fecha_deposito_csv, ID_transaccion_csv]],
        left_on=Usuario_excel,
        right_on=usuario_csv,
        how="left",
        suffixes=("", "_txn")
    )

    # Renombrar y seleccionar columnas finales
    df_resultado = df_resultado.rename(columns={
        fecha_creacion_excel: "Fecha de registro",
        fecha_deposito_csv: "Fecha primer dep√≥sito",
        ID_transaccion_csv: "ID primer dep√≥sito"
    })

    final_cols = [fecha_creacion_excel, ID_usuario_excel, Usuario_excel, fecha_deposito_csv, ID_transaccion_csv]
    final_cols = [c for c in final_cols if c in df_resultado.columns]
    df_resultado = df_resultado[final_cols]

    # Asegurar que las fechas est√©n en el formato correcto para guardar en SQLite
    if fecha_creacion_excel in df_resultado.columns:
        df_resultado[fecha_creacion_excel] = pd.to_datetime(df_resultado[fecha_creacion_excel], errors="coerce").dt.strftime("%Y-%m-%d")
    if fecha_deposito_csv in df_resultado.columns:
        df_resultado[fecha_deposito_csv] = pd.to_datetime(df_resultado[fecha_deposito_csv], errors="coerce").dt.strftime("%Y-%m-%d")

    ### ---------- Guardar en SQLite ---------- ###
    def guardar_en_sqlite(df, nombre_tabla, archivo_db, if_exists="replace"):
        """
        Guarda o reemplaza los datos en la tabla de resultados.
        """
        conn = sqlite3.connect(archivo_db)
        
        # Aseguramos que la tabla exista y est√© vac√≠a si vamos a reemplazar
        df.head(0).to_sql(nombre_tabla, conn, if_exists="append", index=False)
        
        # Convertimos los tipos de datos para evitar errores
        for col in df.columns:
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime("%Y-%m-%d")
            if df[col].dtype == "object":
                df[col] = df[col].astype(str)

        # Reemplazar la tabla completa para evitar duplicados
        df.to_sql(nombre_tabla, conn, if_exists=if_exists, index=False)
        print(f"‚úÖ Tabla de resultados '{nombre_tabla}' actualizada con √©xito.")
        conn.close()

    ### ---------- Llamada a la funcion para guardar en SQLite ---------- ###
    guardar_en_sqlite(df_resultado, name_tabla_resultados, archivo_base_datos, if_exists="replace")
    print("\nüöÄ El proceso ha finalizado correctamente. La tabla de resultados est√° actualizada.")

    ### ---------- Funcion para crear la tabla de resumen diario de registros y FTD's ---------- ###
    def crear_tabla_resumen(conn, name_tabla_resumen):
        """
        Crea una tabla de resumen diario a partir de las tablas de registros y cruce existentes.
        """
        print(f"\n‚è≥ Creando la tabla de resumen diario '{name_tabla_resumen}'...")
        
        # 1. Consulta para contar registros por fecha
        query_registros = "SELECT \"Fecha de creaci√≥n\" AS Fecha, COUNT(*) AS Registros FROM tabla_registros GROUP BY Fecha"
        df_registros_diarios = pd.read_sql(query_registros, conn)
        
        # 2. Consulta para contar primeros dep√≥sitos por fecha
        query_depositos = "SELECT \"Fecha primer dep√≥sito\" AS Fecha, COUNT(\"ID primer dep√≥sito\") AS `Primeros Dep√≥sitos` FROM cruce_registros_transacciones WHERE \"Fecha primer dep√≥sito\" IS NOT NULL GROUP BY Fecha"
        df_depositos_diarios = pd.read_sql(query_depositos, conn)
        
        # 3. Unir ambos DataFrames por la columna de Fecha
        df_resumen = df_registros_diarios.merge(df_depositos_diarios, on="Fecha", how="left")
        
        # 4. Rellenar los valores nulos con cero y convertir a entero
        df_resumen['Primeros Dep√≥sitos'] = df_resumen['Primeros Dep√≥sitos'].fillna(0).astype(int)
        
        # 5. Guardar la tabla de resumen en la base de datos
        df_resumen.to_sql(name_tabla_resumen, conn, if_exists="replace", index=False)
        
        print(f"‚úÖ Tabla de resumen diario '{name_tabla_resumen}' creada y actualizada con √©xito.")
        print("\nüìå Vista previa de la tabla de resumen:")
        print(df_resumen.head())

    # Llamada a la nueva funci√≥n para crear la tabla de resumen
    crear_tabla_resumen(conn, name_tabla_resumen)

    ### ---------- Logica para exportar la tabla resumen a un archivo excel ---------- ###
    def exportar_a_excel(conn, tablas_a_exportar, archivo_excel_salida):
        """
        Exporta m√∫ltiples tablas de la base de datos a un solo archivo de Excel,
        aplicando un estilo de tabla autom√°tico a cada hoja.
        """
        print(f"\n‚è≥ Exportando datos a {archivo_excel_salida} y aplicando estilos de tabla...")
        
        # Crea un nuevo libro de trabajo de openpyxl
        wb = Workbook()
        
        # Elimina la hoja predeterminada 'Sheet'
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        for nombre_tabla in tablas_a_exportar:
            print(f"‚úÖ La tabla '{nombre_tabla}' ha sido exportada a la hoja '{nombre_tabla}'.")
            
            # Lee la tabla SQL en un DataFrame de pandas
            df = pd.read_sql(f"SELECT * FROM {nombre_tabla}", conn)

            # üîπ Nuevo c√≥digo: Convertir columnas de ID a string para evitar notaci√≥n cient√≠fica
            for col in df.columns:
                if 'ID' in col or 'id' in col:
                    df[col] = df[col].astype(str).str.replace(r'\.0$', '', regex=True)

            df = df.replace('nan', '', regex=True)
            df = df.fillna('')
            
            # Crea una nueva hoja con el nombre de la tabla
            ws = wb.create_sheet(nombre_tabla)

            # Inserta los datos del DataFrame en la hoja
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
                    
            # Define el rango de la tabla y el estilo
            num_filas = df.shape[0] + 1
            num_columnas = df.shape[1]
            letra_final = get_column_letter(num_columnas)
            rango_tabla = f"A1:{letra_final}{num_filas}"
            
            # Crea y aplica el estilo de tabla
            tabla = Table(displayName=f"Tabla_{nombre_tabla}", ref=rango_tabla)
            estilo = TableStyleInfo(name="TableStyleMedium9", 
                                    showFirstColumn=False,
                                    showLastColumn=False, 
                                    showRowStripes=True, 
                                    showColumnStripes=False)
            tabla.tableStyleInfo = estilo
            ws.add_table(tabla)

        # Guarda el archivo final
        wb.save(archivo_excel_salida)
        print("\nüöÄ El proceso de exportaci√≥n a Excel ha finalizado correctamente.")

    # Llama a la nueva funci√≥n
    tablas_a_exportar = [name_tabla_resultados, name_tabla_resumen]
    exportar_a_excel(conn, tablas_a_exportar, "Algoritmo_ftd_aciertala_peru.xlsx")

    # Y mueve esta l√≠nea al final del script, despu√©s de todo
    conn.close()

if __name__ == "__main__":
    ad.data_auto_descarga()
    reporte_diario_aciertala()